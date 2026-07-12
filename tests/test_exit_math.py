"""Exit-math generator tests. Verifies the workbook is produced with the
expected structural shape per method; we don't evaluate formulas in tests
(openpyxl can't compute them — that requires Excel/LibreOffice)."""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from angel_memos.exit_math import generate_exit_math
from angel_memos.models import (
    AngelListMetadata,
    Conviction,
    Decision,
    Stage,
    ValuationMethod,
    Verdict,
)


def _angellist() -> AngelListMetadata:
    return AngelListMetadata.model_validate(
        {
            "company": "Spot AI",
            "round_label": "Series B+",
            "stage": Stage.SERIES_B.value,
            "instrument": "Equity",
            "estimated_round_size_usd": 20_000_000,
            "share_class": "Preferred",
            "pre_money_usd": 175_000_000,
            "allocation_usd": 63_000,
            "estimated_expenses_pct": 0.111,
            "leads_investment_usd": 10_000,
            "gross_carry_pct": 0.20,
            "min_investment_usd": 2_000,
            "deadline": date(2025, 11, 20),
            "markets": ["AI", "ML"],
            "founders": ["Rish Gupta", "Sud Bhatija", "Tanuj Thapliyal"],
            "co_investors": ["Scale Venture Partners"],
            "total_prior_capital_usd": 93_000_000,
        }
    )


def _arr_decision() -> Decision:
    return Decision.model_validate(
        {
            "company": "Spot AI",
            "verdict": Verdict.BUY.value,
            "conviction": Conviction.HIGH.value,
            "check_usd": 10_000,
            "post_money_usd": 195_000_000,
            "valuation_method": ValuationMethod.ARR_MULTIPLE.value,
            "current_base_metric_usd": 23_000_000,
            "scenarios": [
                {
                    "name": "Base",
                    "probability": 0.3,
                    "future_dilution": 0.35,
                    "cagr": 0.45,
                    "exit_multiple": 10.0,
                },
                {
                    "name": "Aggressive",
                    "probability": 0.15,
                    "future_dilution": 0.3,
                    "cagr": 0.6,
                    "exit_multiple": 15.0,
                },
                {
                    "name": "Conservative",
                    "probability": 0.25,
                    "future_dilution": 0.45,
                    "cagr": 0.3,
                    "exit_multiple": 7.0,
                },
                {
                    "name": "Home Run",
                    "probability": 0.1,
                    "future_dilution": 0.25,
                    "cagr": 0.75,
                    "exit_multiple": 20.0,
                },
                {
                    "name": "Downside",
                    "probability": 0.2,
                    "future_dilution": 0.6,
                    "cagr": 0.1,
                    "exit_multiple": 4.0,
                },
            ],
            "benchmarks": [
                {
                    "rank_label": "Top 1",
                    "comparable": "Datadog",
                    "terminal_arr_usd": 2_500_000_000,
                    "exit_multiple": 16.0,
                    "exit_valuation_usd": 40_000_000_000,
                },
                {
                    "rank_label": "Top 5",
                    "comparable": "Snowflake",
                    "terminal_arr_usd": 2_000_000_000,
                    "exit_multiple": 14.0,
                    "exit_valuation_usd": 28_000_000_000,
                },
                {
                    "rank_label": "Top 20",
                    "comparable": "Sumo Logic",
                    "terminal_arr_usd": 250_000_000,
                    "exit_multiple": 6.0,
                    "exit_valuation_usd": 1_500_000_000,
                },
            ],
            "top_reasons": ["A", "B", "C"],
            "top_risks": ["X", "Y", "Z"],
            "raw_reasoning": "High conviction.",
        }
    )


def _seed_decision() -> Decision:
    return Decision.model_validate(
        {
            "company": "Seed Co",
            "verdict": Verdict.BUY.value,
            "conviction": Conviction.MEDIUM.value,
            "check_usd": 25_000,
            "post_money_usd": 12_000_000,
            "valuation_method": ValuationMethod.SEED_OUTCOME.value,
            "current_base_metric_usd": None,
            "scenarios": [
                {"name": "zero", "probability": 0.5, "future_dilution": 0.4, "exit_value_usd": 0},
                {
                    "name": "acqui_hire",
                    "probability": 0.2,
                    "future_dilution": 0.4,
                    "exit_value_usd": 30_000_000,
                },
                {
                    "name": "modest",
                    "probability": 0.15,
                    "future_dilution": 0.55,
                    "exit_value_usd": 300_000_000,
                },
                {
                    "name": "breakout",
                    "probability": 0.1,
                    "future_dilution": 0.65,
                    "exit_value_usd": 3_000_000_000,
                },
                {
                    "name": "generational",
                    "probability": 0.05,
                    "future_dilution": 0.7,
                    "exit_value_usd": 30_000_000_000,
                },
            ],
            "benchmarks": [
                {
                    "rank_label": "Top 1",
                    "comparable": "Datadog",
                    "exit_valuation_usd": 40_000_000_000,
                },
                {
                    "rank_label": "Top 5",
                    "comparable": "Sumo Logic",
                    "exit_valuation_usd": 1_500_000_000,
                },
                {
                    "rank_label": "Top 20",
                    "comparable": "Acqui-hire",
                    "exit_valuation_usd": 30_000_000,
                },
            ],
            "top_reasons": ["A", "B", "C"],
            "top_risks": ["X", "Y", "Z"],
            "raw_reasoning": "Power-law bet.",
        }
    )


def test_arr_workbook_creates_file_with_expected_top_inputs(tmp_path: Path) -> None:
    out = tmp_path / "exit_math.xlsx"
    generate_exit_math(_arr_decision(), _angellist(), out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # Top inputs map to the AL fees/carry and decision's post-money + base metric.
    assert ws["B2"].value == pytest.approx(0.111)
    assert ws["B3"].value == pytest.approx(0.20)
    assert ws["B4"].value == 5  # horizon
    assert ws["B5"].value == 195_000_000
    assert ws["B6"].value == 23_000_000


def test_arr_workbook_has_one_row_per_scenario(tmp_path: Path) -> None:
    out = tmp_path / "exit_math.xlsx"
    decision = _arr_decision()
    generate_exit_math(decision, _angellist(), out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # Scenarios live after benchmarks (3 of them). Block lives at row 15 onward.
    scenario_names_col_a = [ws.cell(row=r, column=1).value for r in range(15, 30)]
    expected = [s.name for s in (decision.scenarios or [])]
    for name in expected:
        assert name in scenario_names_col_a


def test_arr_workbook_writes_formulas_for_net_mom(tmp_path: Path) -> None:
    """Net MoM column contains formulas (not bare numbers); we don't evaluate
    them — Excel/LibreOffice will when the user opens the file."""
    out = tmp_path / "exit_math.xlsx"
    generate_exit_math(_arr_decision(), _angellist(), out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # Find the Scenarios block by label, then check Net MoM (column H) on
    # the first data row (one row past the header).
    scenarios_row = next(r for r in range(1, 40) if ws.cell(row=r, column=1).value == "Scenarios")
    first_data_row = scenarios_row + 2
    net_mom_cell = ws.cell(row=first_data_row, column=8).value
    assert isinstance(net_mom_cell, str)
    assert net_mom_cell.startswith("=")
    assert "MAX" in net_mom_cell  # carry waterfall present


def test_seed_workbook_creates_file(tmp_path: Path) -> None:
    out = tmp_path / "seed_exit_math.xlsx"
    generate_exit_math(_seed_decision(), _angellist(), out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # Seed sheet has the directional disclaimer in cell A5.
    assert "DIRECTIONAL" in str(ws["A5"].value)


def test_seed_workbook_has_no_irr_column(tmp_path: Path) -> None:
    """Per user's request: seed exit math shows MoM range, no IRR cell."""
    out = tmp_path / "seed_exit_math.xlsx"
    generate_exit_math(_seed_decision(), _angellist(), out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # Scenario header row should have 5 columns (no IRR header).
    # Find the "Scenarios" label, then check the header row.
    scenarios_row = next(r for r in range(1, 30) if ws.cell(row=r, column=1).value == "Scenarios")
    header_row = scenarios_row + 1
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 7)]
    assert "Net MoM" in headers
    assert "Net IRR" not in headers


def test_exit_math_rejects_pass_decision(tmp_path: Path) -> None:
    decision = Decision.model_validate(
        {
            "company": "Passed Co",
            "verdict": Verdict.PASS.value,
            "conviction": Conviction.LOW.value,
            "check_usd": 0,
            "post_money_usd": 10_000_000,
            "valuation_method": ValuationMethod.ARR_MULTIPLE.value,
            "current_base_metric_usd": None,
            "scenarios": None,
            "benchmarks": None,
            "top_reasons": ["A", "B", "C"],
            "top_risks": ["X", "Y", "Z"],
            "raw_reasoning": "Pass.",
        }
    )
    with pytest.raises(ValueError, match="pass"):
        generate_exit_math(decision, _angellist(), tmp_path / "out.xlsx")


def test_exit_math_rejects_custom_method(tmp_path: Path) -> None:
    decision = Decision.model_validate(
        {
            "company": "Custom Co",
            "verdict": Verdict.BUY.value,
            "conviction": Conviction.HIGH.value,
            "check_usd": 25_000,
            "post_money_usd": 30_000_000,
            "valuation_method": ValuationMethod.CUSTOM.value,
            "current_base_metric_usd": None,
            "scenarios": None,
            "benchmarks": None,
            "top_reasons": ["A", "B", "C"],
            "top_risks": ["X", "Y", "Z"],
            "raw_reasoning": "Bespoke.",
        }
    )
    with pytest.raises(ValueError, match="custom"):
        generate_exit_math(decision, _angellist(), tmp_path / "out.xlsx")
