"""Exit-math xlsx generation.

Programmatically writes the per-deal exit-math workbook from a `Decision` +
`AngelListMetadata`. One sheet per workbook; layout varies by
`valuation_method` because each method has different scenario drivers.

Common structure across methods:
  - Top inputs block: Fees, Carry, Post-money, (Current metric for growth)
  - Benchmarks block: anchors scenarios to named comparables
  - Scenarios block: per-scenario probability + drivers + Net MoM
  - Blended row: SUMPRODUCT of probabilities × Net MoM (and IRR for growth)

The Net MoM formula is universal:
  NetMoM = (1 - fees) / postMoney * (1 - dilution) * exitValuation
         - MAX(0, ((1-fees)/postMoney * (1-dilution) * exitValuation) - 1) * carry

Net IRR (growth methods only, 5-year horizon assumed):
  NetIRR = NetMoM^(1/5) - 1

For seed, IRR is omitted — exit timing is too uncertain to assign a single
horizon, and the user explicitly asked for MoM-only output.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from angel_memos.models import (
    AngelListMetadata,
    ArrBenchmark,
    ArrScenario,
    Decision,
    GmvBenchmark,
    GmvScenario,
    RevenueEbitdaBenchmark,
    RevenueEbitdaScenario,
    RevenuePeBenchmark,
    RevenuePeScenario,
    SeedBenchmark,
    SeedScenario,
    ValuationMethod,
    Verdict,
)

# Growth-stage exit horizon (years). Hard-coded; matches the user's SpotAI
# sheet. If we ever need to parameterize per-deal, add `exit_horizon_years`
# to the Decision schema.
_GROWTH_HORIZON_YEARS = 5

_BOLD = Font(bold=True)


def generate_exit_math(decision: Decision, angellist: AngelListMetadata, out_path: Path) -> Path:
    """Write the per-method exit-math workbook to `out_path`. Returns the
    path on success.

    Raises:
      ValueError: `verdict == pass` or `valuation_method == custom` — these
        cases have no exit math by design.
    """
    if decision.verdict == Verdict.PASS:
        raise ValueError("exit math is omitted for pass decisions")
    if decision.valuation_method == ValuationMethod.CUSTOM:
        raise ValueError("exit math is omitted for custom valuation method")

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover — Workbook() always creates one
        raise RuntimeError("workbook has no active sheet")
    ws.title = "Exit Math"

    method = decision.valuation_method
    match method:
        case ValuationMethod.ARR_MULTIPLE:
            _fill_arr_multiple(ws, decision, angellist)
        case ValuationMethod.REVENUE_EBITDA:
            _fill_revenue_ebitda(ws, decision, angellist)
        case ValuationMethod.REVENUE_PE:
            _fill_revenue_pe(ws, decision, angellist)
        case ValuationMethod.GMV_TAKE:
            _fill_gmv_take(ws, decision, angellist)
        case ValuationMethod.SEED_OUTCOME:
            _fill_seed_outcome(ws, decision, angellist)
        # CUSTOM is unreachable — the guard above raises before the match,
        # and pyright narrows `method` accordingly.

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _write_top_inputs_growth(
    ws: Worksheet,
    angellist: AngelListMetadata,
    decision: Decision,
    base_metric_label: str,
) -> dict[str, str]:
    """Top input block for growth methods. Returns named-cell references."""
    ws["A2"] = "Fees"
    ws["B2"] = angellist.estimated_expenses_pct
    ws["B2"].number_format = "0.0%"

    ws["A3"] = "Carry"
    ws["B3"] = angellist.gross_carry_pct
    ws["B3"].number_format = "0.0%"

    ws["A4"] = "Exit Horizon (Years)"
    ws["B4"] = _GROWTH_HORIZON_YEARS

    ws["A5"] = "Post-Money Entry Valuation"
    ws["B5"] = decision.post_money_usd
    ws["B5"].number_format = '"$"#,##0'

    ws["A6"] = f"Current {base_metric_label}"
    ws["B6"] = decision.current_base_metric_usd or 0
    ws["B6"].number_format = '"$"#,##0'

    for row in range(2, 7):
        ws[f"A{row}"].font = _BOLD

    return {
        "fees": "$B$2",
        "carry": "$B$3",
        "horizon": "$B$4",
        "post_money": "$B$5",
        "base_metric": "$B$6",
    }


def _write_top_inputs_seed(
    ws: Worksheet, angellist: AngelListMetadata, decision: Decision
) -> dict[str, str]:
    """Top input block for seed_outcome — no horizon (outcome-based) and no
    base metric (we don't extrapolate from current numbers)."""
    ws["A2"] = "Fees"
    ws["B2"] = angellist.estimated_expenses_pct
    ws["B2"].number_format = "0.0%"

    ws["A3"] = "Carry"
    ws["B3"] = angellist.gross_carry_pct
    ws["B3"].number_format = "0.0%"

    ws["A4"] = "Post-Money Entry Valuation"
    ws["B4"] = decision.post_money_usd
    ws["B4"].number_format = '"$"#,##0'

    ws["A5"] = "DIRECTIONAL — power-law category, point estimates not load-bearing"
    ws["A5"].font = Font(italic=True)

    for row in range(2, 5):
        ws[f"A{row}"].font = _BOLD

    return {
        "fees": "$B$2",
        "carry": "$B$3",
        "post_money": "$B$4",
    }


def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _BOLD


def _net_mom_formula(
    fees: str, post_money: str, dilution_cell: str, exit_val_cell: str, carry: str
) -> str:
    """Universal Net MoM formula given absolute references to top inputs and
    relative refs to per-row dilution + exit valuation."""
    gross = f"((1-{fees})/{post_money})*(1-{dilution_cell})*{exit_val_cell}"
    return f"={gross} - MAX(0, ({gross}) - 1) * {carry}"


def _net_irr_formula(net_mom_cell: str, horizon: str) -> str:
    """Annualized IRR from Net MoM over the growth horizon."""
    return f"=IFERROR({net_mom_cell}^(1/{horizon})-1, 0)"


def _autosize_columns(ws: Worksheet, max_col: int) -> None:
    """Rough column-width fit. Not precise (openpyxl can't measure text), but
    keeps the generated sheet from looking cramped."""
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 12
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, 40)


# ---------------------------------------------------------------------------
# Method-specific fillers
# ---------------------------------------------------------------------------


def _fill_arr_multiple(ws: Worksheet, decision: Decision, angellist: AngelListMetadata) -> None:
    refs = _write_top_inputs_growth(ws, angellist, decision, "ARR")
    benchmarks = decision.benchmarks or []

    bm_start = 8
    ws.cell(row=bm_start, column=1, value="Benchmarks").font = _BOLD
    _write_header_row(
        ws,
        bm_start + 1,
        ["Rank", "Comparable", "Terminal ARR", "Exit Multiple", "Exit Valuation"],
    )
    for i, b in enumerate(benchmarks):
        if not isinstance(b, ArrBenchmark):
            continue
        r = bm_start + 2 + i
        ws.cell(row=r, column=1, value=b.rank_label)
        ws.cell(row=r, column=2, value=b.comparable)
        ws.cell(row=r, column=3, value=b.terminal_arr_usd).number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=b.exit_multiple).number_format = "0.0"
        ws.cell(row=r, column=5, value=b.exit_valuation_usd).number_format = '"$"#,##0'

    sc_start = bm_start + 4 + len(benchmarks)
    ws.cell(row=sc_start, column=1, value="Scenarios").font = _BOLD
    headers = [
        "Name",
        "Probability",
        "ARR CAGR",
        "Terminal ARR",
        "Exit Multiple",
        "Future Dilution",
        "Exit Valuation",
        "Net MoM",
        "Net IRR",
    ]
    _write_header_row(ws, sc_start + 1, headers)

    scenarios = decision.scenarios or []
    first_data_row = sc_start + 2
    for i, s in enumerate(scenarios):
        if not isinstance(s, ArrScenario):
            continue
        r = first_data_row + i
        ws.cell(row=r, column=1, value=s.name)
        ws.cell(row=r, column=2, value=s.probability).number_format = "0.0%"
        ws.cell(row=r, column=3, value=s.cagr).number_format = "0.0%"
        ws.cell(
            row=r,
            column=4,
            value=f"={refs['base_metric']}*(1+C{r})^{refs['horizon']}",
        ).number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=s.exit_multiple).number_format = "0.0"
        ws.cell(row=r, column=6, value=s.future_dilution).number_format = "0.0%"
        ws.cell(row=r, column=7, value=f"=D{r}*E{r}").number_format = '"$"#,##0'
        ws.cell(
            row=r,
            column=8,
            value=_net_mom_formula(
                refs["fees"], refs["post_money"], f"F{r}", f"G{r}", refs["carry"]
            ),
        ).number_format = "0.00x"
        ws.cell(
            row=r, column=9, value=_net_irr_formula(f"H{r}", refs["horizon"])
        ).number_format = "0.0%"

    blended_row = first_data_row + len(scenarios) + 1
    ws.cell(row=blended_row, column=7, value="Blended").font = _BOLD
    ws.cell(
        row=blended_row,
        column=8,
        value=f"=SUMPRODUCT(B{first_data_row}:B{first_data_row + len(scenarios) - 1},"
        f"H{first_data_row}:H{first_data_row + len(scenarios) - 1})",
    ).number_format = "0.00x"
    ws.cell(
        row=blended_row,
        column=9,
        value=f"=SUMPRODUCT(B{first_data_row}:B{first_data_row + len(scenarios) - 1},"
        f"I{first_data_row}:I{first_data_row + len(scenarios) - 1})",
    ).number_format = "0.0%"

    _autosize_columns(ws, 9)


def _fill_revenue_ebitda(ws: Worksheet, decision: Decision, angellist: AngelListMetadata) -> None:
    refs = _write_top_inputs_growth(ws, angellist, decision, "Revenue")
    benchmarks = decision.benchmarks or []

    bm_start = 8
    ws.cell(row=bm_start, column=1, value="Benchmarks").font = _BOLD
    _write_header_row(
        ws,
        bm_start + 1,
        ["Rank", "Comparable", "Terminal Revenue", "EBITDA Margin", "EV/EBITDA", "Exit Valuation"],
    )
    for i, b in enumerate(benchmarks):
        if not isinstance(b, RevenueEbitdaBenchmark):
            continue
        r = bm_start + 2 + i
        ws.cell(row=r, column=1, value=b.rank_label)
        ws.cell(row=r, column=2, value=b.comparable)
        ws.cell(row=r, column=3, value=b.terminal_revenue_usd).number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=b.ebitda_margin).number_format = "0.0%"
        ws.cell(row=r, column=5, value=b.ev_ebitda).number_format = "0.0"
        ws.cell(row=r, column=6, value=b.exit_valuation_usd).number_format = '"$"#,##0'

    sc_start = bm_start + 4 + len(benchmarks)
    ws.cell(row=sc_start, column=1, value="Scenarios").font = _BOLD
    headers = [
        "Name",
        "Probability",
        "Revenue CAGR",
        "Terminal Revenue",
        "EBITDA Margin",
        "EV/EBITDA",
        "Future Dilution",
        "Exit Valuation",
        "Net MoM",
        "Net IRR",
    ]
    _write_header_row(ws, sc_start + 1, headers)

    scenarios = decision.scenarios or []
    first_data_row = sc_start + 2
    for i, s in enumerate(scenarios):
        if not isinstance(s, RevenueEbitdaScenario):
            continue
        r = first_data_row + i
        ws.cell(row=r, column=1, value=s.name)
        ws.cell(row=r, column=2, value=s.probability).number_format = "0.0%"
        ws.cell(row=r, column=3, value=s.revenue_cagr).number_format = "0.0%"
        ws.cell(
            row=r,
            column=4,
            value=f"={refs['base_metric']}*(1+C{r})^{refs['horizon']}",
        ).number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=s.terminal_ebitda_margin).number_format = "0.0%"
        ws.cell(row=r, column=6, value=s.ev_ebitda).number_format = "0.0"
        ws.cell(row=r, column=7, value=s.future_dilution).number_format = "0.0%"
        ws.cell(row=r, column=8, value=f"=D{r}*E{r}*F{r}").number_format = '"$"#,##0'
        ws.cell(
            row=r,
            column=9,
            value=_net_mom_formula(
                refs["fees"], refs["post_money"], f"G{r}", f"H{r}", refs["carry"]
            ),
        ).number_format = "0.00x"
        ws.cell(
            row=r, column=10, value=_net_irr_formula(f"I{r}", refs["horizon"])
        ).number_format = "0.0%"

    blended_row = first_data_row + len(scenarios) + 1
    ws.cell(row=blended_row, column=8, value="Blended").font = _BOLD
    last = first_data_row + len(scenarios) - 1
    ws.cell(
        row=blended_row,
        column=9,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},I{first_data_row}:I{last})",
    ).number_format = "0.00x"
    ws.cell(
        row=blended_row,
        column=10,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},J{first_data_row}:J{last})",
    ).number_format = "0.0%"

    _autosize_columns(ws, 10)


def _fill_revenue_pe(ws: Worksheet, decision: Decision, angellist: AngelListMetadata) -> None:
    refs = _write_top_inputs_growth(ws, angellist, decision, "Revenue")
    benchmarks = decision.benchmarks or []

    bm_start = 8
    ws.cell(row=bm_start, column=1, value="Benchmarks").font = _BOLD
    _write_header_row(
        ws,
        bm_start + 1,
        ["Rank", "Comparable", "Terminal Revenue", "Net Margin", "P/E", "Exit Valuation"],
    )
    for i, b in enumerate(benchmarks):
        if not isinstance(b, RevenuePeBenchmark):
            continue
        r = bm_start + 2 + i
        ws.cell(row=r, column=1, value=b.rank_label)
        ws.cell(row=r, column=2, value=b.comparable)
        ws.cell(row=r, column=3, value=b.terminal_revenue_usd).number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=b.net_margin).number_format = "0.0%"
        ws.cell(row=r, column=5, value=b.pe_ratio).number_format = "0.0"
        ws.cell(row=r, column=6, value=b.exit_valuation_usd).number_format = '"$"#,##0'

    sc_start = bm_start + 4 + len(benchmarks)
    ws.cell(row=sc_start, column=1, value="Scenarios").font = _BOLD
    headers = [
        "Name",
        "Probability",
        "Revenue CAGR",
        "Terminal Revenue",
        "Net Margin",
        "P/E",
        "Future Dilution",
        "Exit Valuation",
        "Net MoM",
        "Net IRR",
    ]
    _write_header_row(ws, sc_start + 1, headers)

    scenarios = decision.scenarios or []
    first_data_row = sc_start + 2
    for i, s in enumerate(scenarios):
        if not isinstance(s, RevenuePeScenario):
            continue
        r = first_data_row + i
        ws.cell(row=r, column=1, value=s.name)
        ws.cell(row=r, column=2, value=s.probability).number_format = "0.0%"
        ws.cell(row=r, column=3, value=s.revenue_cagr).number_format = "0.0%"
        ws.cell(
            row=r,
            column=4,
            value=f"={refs['base_metric']}*(1+C{r})^{refs['horizon']}",
        ).number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=s.terminal_net_margin).number_format = "0.0%"
        ws.cell(row=r, column=6, value=s.pe_ratio).number_format = "0.0"
        ws.cell(row=r, column=7, value=s.future_dilution).number_format = "0.0%"
        ws.cell(row=r, column=8, value=f"=D{r}*E{r}*F{r}").number_format = '"$"#,##0'
        ws.cell(
            row=r,
            column=9,
            value=_net_mom_formula(
                refs["fees"], refs["post_money"], f"G{r}", f"H{r}", refs["carry"]
            ),
        ).number_format = "0.00x"
        ws.cell(
            row=r, column=10, value=_net_irr_formula(f"I{r}", refs["horizon"])
        ).number_format = "0.0%"

    blended_row = first_data_row + len(scenarios) + 1
    ws.cell(row=blended_row, column=8, value="Blended").font = _BOLD
    last = first_data_row + len(scenarios) - 1
    ws.cell(
        row=blended_row,
        column=9,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},I{first_data_row}:I{last})",
    ).number_format = "0.00x"
    ws.cell(
        row=blended_row,
        column=10,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},J{first_data_row}:J{last})",
    ).number_format = "0.0%"

    _autosize_columns(ws, 10)


def _fill_gmv_take(ws: Worksheet, decision: Decision, angellist: AngelListMetadata) -> None:
    refs = _write_top_inputs_growth(ws, angellist, decision, "GMV")
    benchmarks = decision.benchmarks or []

    bm_start = 8
    ws.cell(row=bm_start, column=1, value="Benchmarks").font = _BOLD
    _write_header_row(
        ws,
        bm_start + 1,
        ["Rank", "Comparable", "Terminal GMV", "Take Rate", "Revenue Multiple", "Exit Valuation"],
    )
    for i, b in enumerate(benchmarks):
        if not isinstance(b, GmvBenchmark):
            continue
        r = bm_start + 2 + i
        ws.cell(row=r, column=1, value=b.rank_label)
        ws.cell(row=r, column=2, value=b.comparable)
        ws.cell(row=r, column=3, value=b.terminal_gmv_usd).number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=b.take_rate).number_format = "0.0%"
        ws.cell(row=r, column=5, value=b.revenue_multiple).number_format = "0.0"
        ws.cell(row=r, column=6, value=b.exit_valuation_usd).number_format = '"$"#,##0'

    sc_start = bm_start + 4 + len(benchmarks)
    ws.cell(row=sc_start, column=1, value="Scenarios").font = _BOLD
    headers = [
        "Name",
        "Probability",
        "GMV CAGR",
        "Terminal GMV",
        "Take Rate",
        "Revenue Multiple",
        "Future Dilution",
        "Exit Valuation",
        "Net MoM",
        "Net IRR",
    ]
    _write_header_row(ws, sc_start + 1, headers)

    scenarios = decision.scenarios or []
    first_data_row = sc_start + 2
    for i, s in enumerate(scenarios):
        if not isinstance(s, GmvScenario):
            continue
        r = first_data_row + i
        ws.cell(row=r, column=1, value=s.name)
        ws.cell(row=r, column=2, value=s.probability).number_format = "0.0%"
        ws.cell(row=r, column=3, value=s.gmv_cagr).number_format = "0.0%"
        ws.cell(
            row=r,
            column=4,
            value=f"={refs['base_metric']}*(1+C{r})^{refs['horizon']}",
        ).number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=s.take_rate).number_format = "0.0%"
        ws.cell(row=r, column=6, value=s.revenue_multiple).number_format = "0.0"
        ws.cell(row=r, column=7, value=s.future_dilution).number_format = "0.0%"
        ws.cell(row=r, column=8, value=f"=D{r}*E{r}*F{r}").number_format = '"$"#,##0'
        ws.cell(
            row=r,
            column=9,
            value=_net_mom_formula(
                refs["fees"], refs["post_money"], f"G{r}", f"H{r}", refs["carry"]
            ),
        ).number_format = "0.00x"
        ws.cell(
            row=r, column=10, value=_net_irr_formula(f"I{r}", refs["horizon"])
        ).number_format = "0.0%"

    blended_row = first_data_row + len(scenarios) + 1
    ws.cell(row=blended_row, column=8, value="Blended").font = _BOLD
    last = first_data_row + len(scenarios) - 1
    ws.cell(
        row=blended_row,
        column=9,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},I{first_data_row}:I{last})",
    ).number_format = "0.00x"
    ws.cell(
        row=blended_row,
        column=10,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},J{first_data_row}:J{last})",
    ).number_format = "0.0%"

    _autosize_columns(ws, 10)


def _fill_seed_outcome(ws: Worksheet, decision: Decision, angellist: AngelListMetadata) -> None:
    refs = _write_top_inputs_seed(ws, angellist, decision)
    benchmarks = decision.benchmarks or []

    bm_start = 7
    ws.cell(row=bm_start, column=1, value="Benchmarks").font = _BOLD
    _write_header_row(ws, bm_start + 1, ["Rank", "Comparable", "Exit Valuation"])
    for i, b in enumerate(benchmarks):
        if not isinstance(b, SeedBenchmark):
            continue
        r = bm_start + 2 + i
        ws.cell(row=r, column=1, value=b.rank_label)
        ws.cell(row=r, column=2, value=b.comparable)
        ws.cell(row=r, column=3, value=b.exit_valuation_usd).number_format = '"$"#,##0'

    sc_start = bm_start + 4 + len(benchmarks)
    ws.cell(row=sc_start, column=1, value="Scenarios").font = _BOLD
    _write_header_row(
        ws,
        sc_start + 1,
        ["Name", "Probability", "Future Dilution", "Exit Valuation", "Net MoM"],
    )

    scenarios = decision.scenarios or []
    first_data_row = sc_start + 2
    for i, s in enumerate(scenarios):
        if not isinstance(s, SeedScenario):
            continue
        r = first_data_row + i
        ws.cell(row=r, column=1, value=s.name)
        ws.cell(row=r, column=2, value=s.probability).number_format = "0.0%"
        ws.cell(row=r, column=3, value=s.future_dilution).number_format = "0.0%"
        ws.cell(row=r, column=4, value=s.exit_value_usd).number_format = '"$"#,##0'
        ws.cell(
            row=r,
            column=5,
            value=_net_mom_formula(
                refs["fees"], refs["post_money"], f"C{r}", f"D{r}", refs["carry"]
            ),
        ).number_format = "0.00x"

    # MoM range — user explicitly asked for range only (no blended IRR for seed).
    blended_row = first_data_row + len(scenarios) + 1
    ws.cell(row=blended_row, column=4, value="MoM Range").font = _BOLD
    last = first_data_row + len(scenarios) - 1
    ws.cell(
        row=blended_row, column=5, value=f"=MIN(E{first_data_row}:E{last})"
    ).number_format = "0.00x"
    ws.cell(row=blended_row + 1, column=4, value="MoM Max").font = _BOLD
    ws.cell(
        row=blended_row + 1, column=5, value=f"=MAX(E{first_data_row}:E{last})"
    ).number_format = "0.00x"
    ws.cell(row=blended_row + 2, column=4, value="Blended MoM").font = _BOLD
    ws.cell(
        row=blended_row + 2,
        column=5,
        value=f"=SUMPRODUCT(B{first_data_row}:B{last},E{first_data_row}:E{last})",
    ).number_format = "0.00x"

    _autosize_columns(ws, 5)
